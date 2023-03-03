import datetime as dt
import os
import tkinter.filedialog
import numpy as np
import pandas as pd
from openpyxl import load_workbook

t=dt.datetime.today().strftime("%d.%m.%Y")


# otwieranie GUI "grafic user interface"
root =tkinter.Tk()
root.withdraw()

url=root.filename=tkinter.filedialog.askopenfilename(title="Select Pricelist file", filetypes=[("Excel files", "*.xlsx")])

df1 = pd.read_excel(url)


# #wczytanie ekstarktu HUAWEI
print('wczytanie ekstarktu HUAWEI')
df2=pd.read_csv(r'\\frfsdtc01\ML12shared_prod$\BufferTable\AX EXTRACTION\CSV\HUAWEI.csv',delimiter=";", usecols=['ItemId','Designation','SubGroup1','SubGroup2','SubGroup3','SubGroup4','SubGroup5','Customer EDI','Origin'])
print('ekstarkt HUAWEI wczytany')
pd.set_option('display.expand_frame_repr', False)


# #usuwanie znakow specjalnych w SKU (regex-[r]-regular expresion) tutaj usuwane puste znaki space, tab, enter [regex= \s]
df1["PartNumber"]=df1["PartNumber"].str.replace(r'\s',"",regex=True)

# # usuwanie przecinków z description
df1['Description']= df1['Description'].astype(str)
df1['Description']=df1['Description'].str.replace(','," ")
df1['Description'] = df1['Description'].str.replace(r'[|]',' ', regex=True)
df1['Description']=df1['Description'].str.replace('  '," ")







# ********************************  UPD  *******************************************************************************************************************************************************************************

UPD=df1.filter(["PartNumber","Description","Software and Hardware Attributes", "Pack Weight\n (kg) ","Pack Dimension\n (D*W*H mm) ",
                "Net Dimension\n (D*W*H mm) ","Discount Category","Product Line","Product Family","Sub Product Family"])

# #zmiana nazw kolumn
UPD.rename(columns={'PartNumber':'SKU'}, inplace=True)


# ----  wymiary ------
#   GROSS
UPD["Pack Dimension\n (D*W*H mm) "]=UPD["Pack Dimension\n (D*W*H mm) "].str.extract(r'([0-9]{1,5}[*][0-9]{1,5}[*][0-9]{1,5})',expand=True)
UPD["Pack Dimension\n (D*W*H mm) "]=UPD["Pack Dimension\n (D*W*H mm) "].str.replace(r'[*]{2}','*',regex=True)
UPD[['Gross width','Gross Height','Gross Depth']]=UPD["Pack Dimension\n (D*W*H mm) "].str.split(r'[*]', expand=True)

UPD['Gross width'] = UPD['Gross width'].fillna(0)
UPD['Gross Height'] = UPD['Gross Height'].fillna(0)
UPD['Gross Depth'] = UPD['Gross Depth'].fillna(0)
UPD = UPD.astype({'Gross width':'int','Gross Height':'int','Gross Depth':'int'})

UPD['Gross width']=UPD['Gross width']/1000
UPD['Gross Height']=UPD['Gross Height']/1000
UPD['Gross Depth']=UPD['Gross Depth']/1000

#  NET
UPD["Net Dimension\n (D*W*H mm) "]=UPD["Net Dimension\n (D*W*H mm) "].str.extract(r'([0-9]{1,5}[*][0-9]{1,5}[*][0-9]{1,5})',expand=True)
UPD["Net Dimension\n (D*W*H mm) "]=UPD["Net Dimension\n (D*W*H mm) "].str.replace(r'[*]{2}','*',regex=True)
UPD[['Net width','Net Height','Net Depth']]=UPD["Net Dimension\n (D*W*H mm) "].str.split(r'[*]', expand=True)

UPD['Net width'] = UPD['Net width'].fillna(0)
UPD['Net Height'] = UPD['Net Height'].fillna(0)
UPD['Net Depth'] = UPD['Net Depth'].fillna(0)
UPD = UPD.astype({'Net width':'int','Net Height':'int','Net Depth':'int'})

UPD['Net width']=UPD['Net width']/1000
UPD['Net Height']=UPD['Net Height']/1000
UPD['Net Depth']=UPD['Net Depth']/1000

UPD.loc[(UPD['Net width']==0) & (UPD['Gross width']!=0),'Net width']=UPD['Gross width']
UPD.loc[(UPD['Net Height']==0) & (UPD['Gross Height']!=0),'Net Height']=UPD['Gross Height']
UPD.loc[(UPD['Net Depth']==0) & (UPD['Gross Depth']!=0),'Net Depth']=UPD['Gross Depth']


UPD.loc[(UPD['Gross width']==0) & (UPD['Net width']!=0),'Gross width']=UPD['Net width']
UPD.loc[(UPD['Gross Height']==0) & (UPD['Net Height']!=0),'Gross Height']=UPD['Net Height']
UPD.loc[(UPD['Gross Depth']==0) & (UPD['Net Depth']!=0),'Gross Depth']=UPD['Net Depth']
#-------------------------


UPD["Pack Weight\n (kg) "] = UPD["Pack Weight\n (kg) "].fillna(0)
UPD.loc[(UPD["Pack Weight\n (kg) "]!=0) | ((UPD['Gross width']!=0)&(UPD['Gross Height']!=0)&(UPD['Gross Depth']!=0))|((UPD['Net width']!=0)&(UPD['Net Height']!=0)&(UPD['Net Depth']!=0)),'W&D']='HARD'



# # ------------  Activity 1
# # HARD
UPD.loc[(UPD['W&D']=='HARD'),'Activity 1']='HARD'
UPD.loc[(UPD["Software and Hardware Attributes"]=='Hardware'),'Activity 1']='HARD'
UPD.loc[(UPD["Discount Category"]=='Hardware'),'Activity 1']='HARD'
#
UPD.loc[(UPD['Description'].str.contains('support',case=False)&(UPD['W&D']=='HARD')),'Activity 1']='HARD +Support'
UPD.loc[(UPD['Description'].str.contains('Power Suply',case=False))& (~UPD['Description'].str.contains('with',case=False)& (UPD['W&D']=='HARD')),'Activity 1']='HARD Power Suply'

# # SOFT
UPD.loc[(UPD['Software and Hardware Attributes']=='Self-developed software')|((UPD['Software and Hardware Attributes']=='Software Annuity')),'Activity 1']='SOFT Licences'
UPD.loc[(UPD['Description'].str.contains('upgrade',case=False)&(UPD["Activity 1"]=='SOFT Licences')),'Activity 1']='SOFT Upgrade'
UPD.loc[UPD['Discount Category'].str.contains("License", case=False,na=False),'Activity 1']='SOFT Licences'

# UPD.loc[(UPD['Description'].str.contains('subscript',case=False)&(UPD["Activity 1"]=='SOFT Licences')),'Activity 1']='SOFT Subscription'
# UPD.loc[(UPD['Description'].str.contains('secur',case=False)&(UPD["Activity 1"]=='SOFT Licences')),'Activity 1']='SOFT SECURE'

# # SERVICE
UPD.loc[UPD["Software and Hardware Attributes"]=='Service','Activity 1']='SERVICE'

UPD.loc[((UPD['Discount Category']=='Outsourcing')& (UPD['W&D']!='HARD')&(UPD['Description'].str.contains('service',case=False))),'Activity 1']='SERVICE'

UPD.loc[UPD['Activity 1'].isnull(),'Activity 1']='SERVICE'


#(UPD["Pack Weight\n (kg) "].notnull())or (UPD["Pack Dimension\n (D*W*H mm) "].notnull()) or (UPD["Net Dimension\n (D*W*H mm) "].notnull()))&


# UPD.loc[(UPD["Method of Delivery"]=='Electronic') & ((UPD['Description'].str.contains('service',case=False)) | (UPD['Description'].str.contains('support',case=False))),'Activity 1']='SERVICE'


# # ------------     Activity 2    --------------

UPD.loc[UPD['Activity 1']=='HARD','Activity 2']='OTHERS'
UPD.loc[((UPD['Activity 1']=='HARD')&(UPD['Description'].str.contains('server',case=False))),'Activity 2']='SERVER'
UPD.loc[UPD['Activity 1']=='HARD +Support','Activity 2']='OTHERS'
UPD.loc[UPD['Activity 1']=='SOFT Licences','Activity 2']='OTHERS'
UPD.loc[UPD['Activity 1']=='SOFT Upgrade','Activity 2']='OTHERS'
UPD.loc[UPD['Activity 1']=='SOFT SECURE','Activity 2']='SECURITY'
UPD.loc[UPD['Activity 1']=='SOFT Subscription','Activity 2']='OTHERS'
UPD.loc[UPD['Activity 1']=='SERVICE','Activity 2']='MAINT'
UPD.loc[(UPD['Activity 1']=='SERVICE')&(UPD['Discount Category'].str.contains('Training',case=False))&(UPD['Product Family'].str.contains('Training',case=False)),'Activity 2']='TRAINING'


# # UPD.loc[(UPD['Activity 1']=='SOFT Subscription')&((UPD['Description'].str.contains('secur',case=False))'Activity 2']='OTHERS'
# UPD.loc[(UPD['Description'].str.contains('secur',case=False)&(UPD["Activity 1"]=='SOFT Subscription')),'Activity 2']='SECURITY'
# UPD.loc[UPD['Activity 1']=='SERVICE','Activity 2']='OTHERS'


# # ------------     Activity 3    ---------------

UPD.loc[(UPD["Description"].str.contains('renewal',case=False)) & (UPD['Activity 1']!='HARD'),'Activity 3']='RENEWAL'
UPD.loc[UPD['Activity 3']!='RENEWAL','Activity 3']='INITIAL'

# #dodawanie nowych kolumn i uzupełnianie
UPD.insert(loc=0, column='Item Group', value="HUAWEI")
UPD.insert(loc=2,column="Vendor SKU", value="")
UPD.insert(loc=3, column="Item Type", value='Item')
UPD['Inventory Model Group']='FIFOARW03'
UPD['Life Cycle']='Online'
UPD['Stock Management']='BACK TO BACK'
UPD['Finance Project Category']=UPD['Item Group']
UPD['ItemPrimaryVendId']=''
UPD['Volume']=''
UPD['Legacy Id']=''
UPD['Customer EDi']='YES'
UPD['List Price UpDate']='YES'
UPD['Dual Use']='YES'
UPD['Virtual Item']='NO'
UPD['Arrow Brand']='HUA'
UPD['Purchase Delivery Time']=''
UPD['Sales Delivery Time']=''
UPD['Production Type']='NONE'
UPD['Unit point']=''
UPD['Warranty']=''
UPD['Renewal term']=''
UPD['Origin']='CHN'
UPD.rename(columns={"Pack Weight\n (kg) ":"Weight"}, inplace=True)
UPD['Tare Weight']=UPD['Weight']*0.2152
UPD['Finance Activity']=UPD['Activity 1']



path=os.path.dirname(url)

#-----   przypisywanie SUBGRUPY 1
sub1=pd.read_excel(path+"/subgrup1.xlsx",usecols=['Sub group 1','Description'])
sub1.rename(columns={'Description':'Description_sub1','Sub group 1':'SubGroup1'}, inplace=True)
UPD=pd.merge(UPD,sub1,left_on='Software and Hardware Attributes',right_on='Description_sub1', how='left')
UPD.loc[UPD['Software and Hardware Attributes']==' ','SubGroup1']='####'


#------   przypisywanie SUBGRUPY 2
sub2=pd.read_excel(path+"/subgrup2.xlsx",usecols=['Sub group 2','Description'])
sub2.rename(columns={'Description':'Description_sub2','Sub group 2':'SubGroup2'}, inplace=True)
# #usuwanie znakow specjalnych na końcu i na początku
UPD['Product Family']=UPD['Product Family'].str.strip()
UPD=pd.merge(UPD,sub2,left_on='Product Family',right_on='Description_sub2', how='left')
UPD.loc[UPD['Product Family']==' ','SubGroup2']='####'


#------   przypisywanie SUBGRUPY 3
sub3=pd.read_excel(path+"/subgrup3.xlsx",usecols=['Sub group 3','Description'])
sub3.rename(columns={'Description':'Description_sub3','Sub group 3':'SubGroup3'}, inplace=True)
# #usuwanie znakow specjalnych na końcu i na początku
UPD['Product Line']=UPD['Product Line'].str.strip()
UPD=pd.merge(UPD,sub3,left_on='Product Line',right_on='Description_sub3', how='left')
UPD.loc[UPD['Product Line']==' ','SubGroup3']='####'
UPD.loc[UPD['Product Line']=='','SubGroup3']='####'

#------   przypisywanie SUBGRUPY 4
sub4=pd.read_excel(path+"/subgrup4.xlsx",usecols=['Sub group 4','Description'])
sub4.rename(columns={'Description':'Description_sub4','Sub group 4':'SubGroup4'}, inplace=True)
# #usuwanie znakow specjalnych na końcu i na początku
UPD['Discount Category']=UPD['Discount Category'].str.strip()
UPD=pd.merge(UPD,sub4,left_on='Discount Category',right_on='Description_sub4', how='left')
UPD.loc[UPD['Discount Category']==' ','SubGroup4']='####'
UPD.loc[UPD['Discount Category'].isnull(),'SubGroup4']='####'

#------   przypisywanie SUBGRUPY 5
sub5=pd.read_excel(path+"/subgrup5.xlsx",usecols=['Sub group 5','Description'])
sub5.rename(columns={'Description':'Description_sub5','Sub group 5':'SubGroup5'}, inplace=True)
# #usuwanie znakow specjalnych na końcu i na początku
UPD['Sub Product Family']=UPD['Sub Product Family'].str.strip()
UPD=pd.merge(UPD,sub5,left_on='Sub Product Family',right_on='Description_sub5', how='left')
UPD.loc[UPD['Sub Product Family']==' ','SubGroup5']='####'
UPD.loc[UPD['Sub Product Family']=='','SubGroup5']='####'
UPD.loc[UPD['Sub Product Family'].isnull(),'SubGroup5']='####'


data = [('HARD','SECURITY','INITIAL','84714900','HW','GROSS','HWR'),
        ('HARD', 'SERVER', 'INITIAL','84714100','HW','GROSS','HWR'),
        ('HARD','STORAGE', 'INITIAL','84714900','HW','GROSS','HWR'),
        ('HARD','OTHERS','INITIAL','84714900','HW','GROSS','HWR'),
        ('HARD Cable','OTHERS','INITIAL','85444210','HW','GROSS','HWR'),
        ('HARD Power Suply','OTHERS','INITIAL','85044030','HW','GROSS','HWR'),
        ('HARD +Support','SECURITY','INITIAL','84714900','HW_SVC','GROSS','SBH'),
        ('HARD +Support', 'SERVER', 'INITIAL','84714100','HW_SVC','GROSS','SBH'),
        ('HARD +Support','STORAGE', 'INITIAL','84714900','HW_SVC','GROSS','SBH'),
        ('HARD +Support','OTHERS','INITIAL','84714900','HW_SVC','GROSS','SBH'),


        #  SERVICES  #
        ('SERVICE','CLOUD','INITIAL','00000000','XAAS','NET','SAA'),
        ('SERVICE','CLOUD','RENEWAL','00000000','RNW','NET','SPR'),
        ('SERVICE','INTEG','INITIAL','00000000','SVC_MAINT','NET','SPN'),
        ('SERVICE','INTEG','RENEWAL','00000000','RNW','NET','SPR'),
        ('SERVICE','LOGISTICS','INITIAL','00000000','HW','GROSS','FRT'),
        ('SERVICE','LOGISTICS','RENEWAL','00000000','HW','GROSS','FRT'),
        ('SERVICE','MAINT','INITIAL','00000000','SVC_MAINT','NET','SPN'),
        ('SERVICE','MAINT','RENEWAL','00000000','RNW','NET','SPR'),
        ('SERVICE','OTHERS','INITIAL','00000000','SVC_MAINT','NET','SPN'),
        ('SERVICE','OTHERS','RENEWAL','00000000','RNW','NET','SPR'),
        ('SERVICE', 'SUBSCRIPT', 'INITIAL', '00000000', 'SVC_MAINT', 'NET', 'SPN'),
        ('SERVICE', 'SUBSCRIPT', 'RENEWAL', '00000000', 'RNW', 'NET', 'SPR'),
        ('SERVICE','PROFSERV', 'INITIAL', '00000000', 'SVC_TPP', 'NET', 'VPS'),
        ('SERVICE','PROFSERV', 'RENEWAL', '00000000', 'RNW', 'NET', 'SPR'),
        # TRENING #
        ('SERVICE','TRAINING', 'INITIAL', '00000000', 'SVC_TPP', 'NET', 'WAT'),

        #  SOFT  #
        ('SOFT CLOUD','CLOUD','INITIAL','00000000','XAAS','NET','SAA'),
        ('SOFT SECURE','SECURITY','INITIAL','00000000','SW_AV','NET','SWV'),
        ('SOFT SECURE','SECURITY','RENEWAL','00000000','SW_AV','NET','SWV'),
        ('SOFT Upgrade','OTHERS','INITIAL','00000000','SW_U','GROSS','SWN'),
        ('SOFT Upgrade','SECURITY','INITIAL','00000000','SW_U','GROSS','SWN'),
        ('SOFT Licences','CLOUD','RENEWAL','00000000','SW_R','GROSS','SWR'),
        ('SOFT Licences','OTHERS','INITIAL','00000000','SW_P','GROSS','SWN'),
        ('SOFT Licences','OTHERS','RENEWAL','00000000','SW_R','GROSS','SWR'),
        ('SOFT Licences', 'SECURITY', 'INITIAL', '00000000', 'SW_P', 'GROSS', 'SWN'),
        ('SOFT Licences', 'SECURITY', 'RENEWAL', '00000000', 'SW_R', 'GROSS', 'SWR'),
        ('SOFT Licences', 'STORAGE MG', 'INITIAL', '00000000', 'SW_P', 'GROSS', 'SWN'),
        ('SOFT Licences', 'STORAGE MG', 'RENEWAL', '00000000', 'SW_R', 'GROSS', 'SWR'),
        ('SOFT Licence + Support', 'CLOUD', 'INITIAL', '00000000', 'SWP_SVC', 'GROSS', 'SBS'),
        ('SOFT Licence + Support', 'OTHERS', 'INITIAL', '00000000', 'SWP_SVC', 'GROSS', 'SBS'),
        ('SOFT Licence + Support', 'SECURITY', 'INITIAL', '00000000', 'SWP_SVC', 'GROSS', 'SBS'),
        ('SOFT Licence + Support', 'STORAGE MG', 'INITIAL', '00000000', 'SWP_SVC', 'GROSS', 'SBS'),
        ('SOFT UPD Licence + Support', 'CLOUD', 'INITIAL', '00000000', 'SWU_SVC ', 'GROSS', 'SBS'),
        ('SOFT UPD Licence + Support', 'OTHERS', 'INITIAL', '00000000', 'SWU_SVC ', 'GROSS', 'SBS'),
        ('SOFT UPD Licence + Support', 'SECURITY', 'INITIAL', '00000000', 'SWU_SVC ', 'GROSS', 'SBS'),
        ('SOFT UPD Licence + Support', 'STORAGE MG', 'INITIAL', '00000000', 'SWU_SVC ', 'GROSS', 'SBS'),

        ('SOFT Subscription', 'OTHERS', 'INITIAL', '00000000', 'SW_FT', 'GROSS', 'SWF'),
        ('SOFT Subscription', 'OTHERS', 'RENEWAL', '00000000', 'SW_R', 'GROSS', 'SWR'),
        ('SOFT Subscription', 'SECURITY', 'INITIAL', '00000000', 'SW_FT', 'GROSS', 'SWF'),
        ('SOFT Subscription', 'SECURITY', 'RENEWAL', '00000000', 'SW_R', 'GROSS', 'SWR'),
        ('SOFT Subscription', 'STORAGE MG', 'INITIAL', '00000000', 'SW_FT', 'GROSS', 'SWF'),
        ('SOFT Subscription', 'STORAGE MG', 'RENEWAL', '00000000', 'SW_R', 'GROSS', 'SWR'),
        ('SOFT Subscript +Support', 'CLOUD', 'INITIAL', '00000000', 'SWFT_SVC', 'GROSS', 'SBF'),
        ('SOFT Subscript +Support', 'OTHERS', 'INITIAL', '00000000', 'SWFT_SVC', 'GROSS', 'SBF'),
        ('SOFT Subscript +Support', 'SECURITY', 'INITIAL', '00000000', 'SWFT_SVC', 'GROSS', 'SBF'),
        ('SOFT Subscript +Support', 'STORAGE MG', 'INITIAL', '00000000', 'SWFT_SVC', 'GROSS', 'SBF'),

        #  ELA  #
        ('SOFT ELA', 'OTHERS', 'INITIAL', '00000000', 'ELA', 'NET', 'ELA'),
        ('SOFT ELA', 'OTHERS', 'RENEWAL', '00000000', 'ELA', 'NET', 'ELA'),
        #  HYBRYD  #
        ('SOFT HYBD', 'CLOUD', 'INITIAL', '00000000', 'HYBD', 'NET', 'HYB'),
        ('SOFT HYBD', 'OTHERS', 'INITIAL', '00000000', 'HYBD', 'NET', 'HYB')
        ]


gn = pd.DataFrame.from_records(data, columns =['Activity 1', 'Activity 2', 'Activity 3','Intrastat Code','Gross/Net Classification','Gross/Net','SUBBRAND'])
gn['merge']=gn['Activity 1']+gn['Activity 2']+gn['Activity 3']
gn1=gn[['merge','Intrastat Code','Gross/Net Classification','Gross/Net','SUBBRAND']]


UPD['merge']=UPD['Activity 1']+UPD['Activity 2']+UPD['Activity 3']


UPD2=UPD.merge(gn1,on="merge", how='left')

###  Special VAT Code
UPD2.loc[(UPD2['Activity 1']=='SERVICE')&(UPD2['Activity 2']=='MAINT')&(UPD2['Description'].str.contains('HARD'or 'warranty',case=False)),'Special VAT Code']='SPVAT0001'


# ### Dimension Group
UPD2["Dimension Group"]="PHYSICAL"
UPD2.loc[UPD2["Intrastat Code"]=="00000000","Dimension Group"]="STDBATCH3"

#### Serial Number Group
UPD2['Serial Number Group']="SN-AECS"
serial={'00000000':"SN-AECS",'85444210':"SN-AECS",'85044030':"SN-AECS",}
UPD2['Serial Number Group']=UPD2["Intrastat Code"].map(serial)

UPD2.loc[UPD2['Intrastat Code']=='00000000','Virtual Item']='YES'
UPD2['Activity 1'] = UPD2['Activity 1'].str.extract(r'(HARD|SOFT|SERVICE)')
UPD2['Finance Activity'] = UPD2['Finance Activity'].str.extract(r'(HARD|SOFT|SERVICE)')

###  Special markers

UPD2.loc[  ((UPD2['Activity 1']=='SERVICE')  &  (UPD2['Activity 2']=='TRAINING')) & (~UPD2['Description'].str.contains('|'.join(['fee','Travel costs','Travel expenses','Conference','event','training materials', 'books']),case=False)), 'Special marker'] ="GTU_12"
UPD2.loc[  (UPD2['Activity 1']=='HARD')  &  ((UPD2['Activity 2']=='SERVER')|(UPD2['Activity 2']=='STORAGE')), 'Special marker'] ="MPP_GTU_06"

# sprawdzanie brakujacych subgrup
new_sub=UPD2.copy()

new_sub.loc[new_sub['SubGroup1'].isnull(),'NEW_Subgrup']='NEW_SUB_1'
new_sub.loc[new_sub['NEW_Subgrup']=='NEW_SUB_1','NEW_Subgrup_description']=new_sub['Software and Hardware Attributes']

new_sub.loc[new_sub['SubGroup2'].isnull(),'NEW_Subgrup']='NEW_SUB_2'
new_sub.loc[new_sub['NEW_Subgrup']=='NEW_SUB_2','NEW_Subgrup_description']=new_sub['Product Family']

new_sub.loc[new_sub['SubGroup3'].isnull(),'NEW_Subgrup']='NEW_SUB_3'
new_sub.loc[new_sub['NEW_Subgrup']=='NEW_SUB_3','NEW_Subgrup_description']=new_sub['Product Line']

new_sub.loc[new_sub['SubGroup4'].isnull(),'NEW_Subgrup']='NEW_SUB_4'
new_sub.loc[new_sub['NEW_Subgrup']=='NEW_SUB_4','NEW_Subgrup_description']=new_sub['Discount Category']

new_sub.loc[new_sub['SubGroup5'].isnull(),'NEW_Subgrup']='NEW_SUB_5'
new_sub.loc[new_sub['NEW_Subgrup']=='NEW_SUB_5','NEW_Subgrup_description']=new_sub['Sub Product Family']

new_sub=new_sub[['NEW_Subgrup','NEW_Subgrup_description']]
new_sub=new_sub[new_sub['NEW_Subgrup'].notnull()]
new_sub.to_excel(path+"/HUAWEI-NEW-SUBGRUPS.xlsx",sheet_name='UPD', startrow=1,index=False)



# #porzadkowanie kolumn
UPD2=UPD2[['Item Group','SKU','Vendor SKU','Item Type','Intrastat Code','Dimension Group','Serial Number Group','Inventory Model Group'
         ,'Life Cycle','Activity 1','Activity 2','Activity 3','Stock Management','SubGroup1','SubGroup2','SubGroup3','SubGroup4',
         'SubGroup5','Description','ItemPrimaryVendId','Weight','Tare Weight','Gross width','Gross Height','Gross Depth','Net width','Net Height','Net Depth','Volume','Legacy Id','Finance Project Category',
         'Finance Activity','Customer EDi','List Price UpDate','Dual Use','Virtual Item','Arrow Brand','Gross/Net','Gross/Net Classification',
         'Purchase Delivery Time','Sales Delivery Time','Production Type','Unit point','Warranty','SUBBRAND','Renewal term','Special VAT Code',
        'Origin','Special marker']]


#
roznica=UPD2.loc[~UPD2['SKU'].isin(df2['ItemId'])]
# path=os.path.dirname(url)

if not roznica.empty:

        roznica.to_excel(path+"/HUAWEI-A86227-UPD-"+t+"-SKU.xlsx",sheet_name='UPD', startrow=1,index=False)

        wbook=load_workbook(path+"/HUAWEI-A86227-UPD-"+t+"-SKU.xlsx")
        sheet=wbook.active
        sheet['A1']='iAssetSync'
        sheet['B1']='No'

        wbook.save(path+"/HUAWEI-A86227-UPD-"+t+"-SKU.xlsx")

        print("UPD saved in direction:")
        print(path+"/HUAWEI-A86227-UPD-"+t+"-SKU.xlsx")
else:
        print("UPD file not saved, there is no new items")

UPD.info(memory_usage='deep')





#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# # *************************   TAR   ***************************************************************************************************************************************************************************************

# # #filtrowanie niezbednych danych
TAR=df1.filter(["PartNumber",'List Price\n (EUR)',"Authorization Discount Off",'Authorization Unit Price\n(EUR FOB HongKong)'])


# # #zmiana nazw kolumn
TAR.rename(columns={'PartNumber':'SKU','List Price\n (EUR)':'Public Price','Authorization Discount Off':'StdRebate','Authorization Unit Price\n(EUR FOB HongKong)':'Channel Price'}, inplace=True)
#
# # #dodawanie nowych kolumn i uzupełnianie
TAR.insert(loc=1, column='AccountCode', value="ALL")
TAR.insert(loc=2, column='Currency', value="EUR")
TAR.insert(loc=3, column='Quantity', value="1")
TAR['Valid From']=t
TAR['Valid To']=""
TAR['SearchAgain']="YES"
TAR['UnitID']="PCS0DEC"
TAR['LegalEntity']="101-141-311-501-502-550-560-580-582-700"
TAR['StdRebate']=TAR['StdRebate']*100
TAR['Item Group']='HUAWEI'

# # #porzadkowanie kolumn
TAR=TAR[['SKU', 'AccountCode','LegalEntity', 'Currency', 'Quantity', 'Public Price', 'StdRebate','Channel Price', 'Valid From', 'Valid To', 'SearchAgain', 'UnitID','Item Group']]

# # # ustawianie numeru SKu jako indeks
TAR.set_index('SKU', inplace=True)

# usuwanie SKu z zerowymi cenami i uzupelnianie brakujacych wartość
TAR= TAR[TAR['Public Price']!= 0]
TAR.loc[TAR['Channel Price']==' ','Channel Price']=TAR['Public Price']
TAR['StdRebate']= TAR['StdRebate'].replace(np.nan, 0 )

TAR['Channel Price']=TAR['Channel Price'].astype('float64')


#zapisywanie do excela
TAR.to_excel(path+"/HUAWEI-A86227-TAR-"+t+"-SKU.xlsx",sheet_name='TAR', startrow=1)
# print(df1.columns)

wbook=load_workbook(path+"/HUAWEI-A86227-TAR-"+t+"-SKU.xlsx")
sheet=wbook.active
sheet['A1']='iAssetSync'
sheet['B1']='No'
wbook.save(path+"/HUAWEI-A86227-TAR-"+t+"-SKU.xlsx")

# print(TAR)

print("TAR saved in folder:")
print(path+"/HUAWEI-A86227-TAR-"+t+"-SKU.xlsx")

# TAR.info(memory_usage='deep')


#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# # -------- AMD ***********************************************************************************************************************************************************************************************************

# df2 to ekstrakt Huawei
# df1 to dane z pricelisty

df1["Description"]=df1["Description"].str.strip()
df2['Designation']=df2['Designation'].str.strip()


# # amd_merge=pd.merge(df1,df2,left_on='SKU',right_on='ItemId')

amd_merge=pd.merge(df2,df1,left_on='ItemId',right_on='PartNumber',how= 'left', indicator=True)

amd=amd_merge.copy()


amd.loc[amd['ItemId'].isnull(),'ItemId']=amd['PartNumber']


# Sprawdzanie opisów
amd.loc[amd["Description"]!= amd['Designation'],'correct Description']=amd["Description"]

# pomijanie opisów dluzszych niz 180 znaków
amd["Description"]= amd["Description"].astype(str)
amd.loc[amd["Description"].str.len()> 180,'correct Description']=amd['Designation']

# oczyszczanie description z przecinków
amd['correct Description']=amd['correct Description'].str.replace(',',' ')
amd['correct Description']=amd['correct Description'].str.replace('  ',' ')

#jezeli po oczyszczeniu oppisy sa takie same usuń z 'To correct'
amd.loc[amd['correct Description']== amd['Designation'],'correct Description']=np.NaN


# # uzupełnianie 'Origin'
amd.loc[amd['Origin']!="CHN",'correct Origin']='CHN'


# zmiana 'Customer EDI'
amd.loc[((amd['_merge']=='left_only')&(amd['Customer EDI']=='Yes')),'correct Customer EDI']='NO'
amd.loc[((amd['_merge']=='both')&(amd['Customer EDI']=='No')),'correct Customer EDI']='YES'

# # sprawdzanie subgrup
# amd.loc[(amd['CATEGORY']!=amd['SubGroup4']),'correct SubGroup4']=amd['CATEGORY']
# # 'CATEGORY':'SubGroup4'

amd.loc[(amd['correct Customer EDI'].notnull() | amd['correct Description'].notnull() | amd['correct Origin'].notnull()),'AMD']="To correct"


amd=amd.loc[amd['AMD']=="To correct"]
amd.fillna({'correct Origin':"CHN",'correct Customer EDI':amd['Customer EDI'],'correct Description':amd['Designation']}, inplace=True)

amd.to_excel(path+"/HUAWEI-A86227-TESTtttttt-"+t+"-SKU.xlsx",sheet_name='AMD', startrow=1)

amd2=amd.copy()
amd2.set_index('ItemId', inplace=True)

amd2=amd2[['correct Description','correct Customer EDI','correct Origin']]
amd2.columns = amd2.columns.str.replace(r'correct ', '')

amd2.index.names = ['SKU']



# #zapisywanie do excela
amd2.to_excel(path+"/HUAWEI-A86227-AMD-"+t+"-SKU.xlsx",sheet_name='AMD', startrow=1)

wbook=load_workbook(path+"/HUAWEI-A86227-AMD-"+t+"-SKU.xlsx")
sheet=wbook.active
sheet['A1']='iAssetSync'
sheet['B1']='No'

wbook.save(path+"/HUAWEI-A86227-AMD-"+t+"-SKU.xlsx")

print("AMD saved in folder:")
print(path+"/HUAWEI-A86227-AMD-"+t+"-SKU.xlsx")